#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
CONTRACTS_DIR = ROOT / "contracts"
RAW_SNAPSHOT_ROOT = Path.home() / "workspace/raw/05-Projects/MAformac/source-snapshots"

CORE_SHEETS = ("airControl", "carControl", "cmd")
FOLLOWUP_SHEETS = ("airControl", "carControl")

SOURCE_FILES = [
    {
        "key": "semantic_protocol_edit",
        "filename": "公版语义四级功能协议表-编辑版.xlsx",
        "authority_kind": "c1_primary",
        "expected_sheets": list(CORE_SHEETS),
        "contract_source": True,
    },
    {
        "key": "multilingual_expansion_v1",
        "filename": "多语种公版语义四级功能展开协议表V1版.xlsx",
        "authority_kind": "c1_reference",
        "expected_sheets": list(CORE_SHEETS),
        "contract_source": False,
    },
    {
        "key": "vehicle_function_points",
        "filename": "车控功能打点表.xlsx",
        "authority_kind": "c1_reference",
        "expected_sheets": list(CORE_SHEETS),
        "contract_source": False,
    },
    {
        "key": "followup_function_list",
        "filename": "上下文二次交互功能清单.xlsx",
        "authority_kind": "c1_followup_source",
        "expected_sheets": list(FOLLOWUP_SHEETS),
        "contract_source": False,
    },
]

SEMANTIC_HEADERS = {
    "primary_function": "一级功能",
    "secondary_zh": "二级功能-中文",
    "secondary_en": "二级功能-英文",
    "tertiary_zh": "三级功能-中文",
    "tertiary_en": "三级功能-英文",
    "function_with_value": "四级功能-带协议赋值",
    "description": "功能描述",
    "nlu_protocol": "NLU协议",
    "nlu_range": "NLU协议字段取值范围",
    "function_text": "四级功能",
    "ds_protocol": "四级功能DS协议",
    "semantic_range": "四级功能DS协议字段取值范围",
    "example_utterance": "示例说法",
    "action_code": "功能类型编码",
    "fc_fuzzy": "FC模糊说",
    "fc_free": "FC自由说",
}

FOLLOWUP_HEADERS = {
    "first_zh": "首轮三级功能-中文",
    "first_intent": "首轮三级功能-英文",
    "first_examples": "首轮示例说法",
    "inherited_slots": "可继承内容",
    "second_zh": "次轮支持三级功能-中文",
    "second_intent": "次轮支持三级功能-英文",
    "second_examples": "次轮示例说法",
    "rewrite_policy": "继承逻辑说明",
}

ACTION_PREFIX = re.compile(
    r"^(open|close|switch|toggle|raise|lower|adjust|set|query|increase|decrease|"
    r"activate|deactivate|lock|unlock|pause|resume|start|stop|save|enable|disable|"
    r"enter|exit|reset|select|confirm|cancel|delete|add|search|connect|disconnect|"
    r"fold|unfold|check|move|navigate|play|listen)_"
)
VALUE_SUFFIX = re.compile(
    r"_(by_number|by_exp|by_percent|by_specific_value|to_number|to_exp|to_gear|"
    r"to_max|to_min|to_percent|to_extremum|to_extreme_direction|little|no_value|"
    r"to_value|by_direction|by_gear|to_specific_value)$"
)


class C1Error(RuntimeError):
    pass


class ValueExtractionError(C1Error):
    def __init__(self, message: str, source_sheet: str, source_row_no: int, payload: Any):
        super().__init__(message)
        self.source_sheet = source_sheet
        self.source_row_no = source_row_no
        self.payload = payload


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def downloads_path(filename: str) -> Path:
    return Path.home() / "Downloads" / filename


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            value = int(value)
    text = unicodedata.normalize("NFC", str(value))
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    return "\n".join(lines).strip()


def stable_json(value: Any, *, ensure_ascii: bool = True) -> str:
    return json.dumps(
        value,
        ensure_ascii=ensure_ascii,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_json(value: Any) -> str:
    return sha256_text(stable_json(value, ensure_ascii=False))


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=str(path.parent), delete=False) as tmp:
        tmp.write(content)
        tmp_name = tmp.name
    os.replace(tmp_name, path)


def atomic_write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    content = "".join(stable_json(row) + "\n" for row in rows)
    atomic_write_text(path, content)
    validate_jsonl(path)


def validate_jsonl(path: Path) -> None:
    with path.open("rb") as f:
        first = f.read(3)
        if first == b"\xef\xbb\xbf":
            raise C1Error(f"{path} has a UTF-8 BOM")
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            if not line.strip():
                raise C1Error(f"{path}:{line_no} is blank")
            json.loads(line, object_pairs_hook=reject_duplicate_json_keys)


def reject_duplicate_json_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in pairs:
        if key in out:
            raise ValueError(f"duplicate JSON key: {key}")
        out[key] = value
    return out


def safe_load_yaml(path: Path) -> Any:
    class UniqueKeyLoader(yaml.SafeLoader):
        pass

    def construct_mapping(loader: yaml.Loader, node: yaml.Node, deep: bool = False) -> dict[Any, Any]:
        mapping = {}
        for key_node, value_node in node.value:
            key = loader.construct_object(key_node, deep=deep)
            if key in mapping:
                raise C1Error(f"{path} has duplicate YAML key: {key}")
            mapping[key] = loader.construct_object(value_node, deep=deep)
        return mapping

    UniqueKeyLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        construct_mapping,
    )
    with path.open("r", encoding="utf-8") as f:
        return yaml.load(f, Loader=UniqueKeyLoader)


def dump_yaml(data: Any) -> str:
    return yaml.safe_dump(
        data,
        allow_unicode=False,
        sort_keys=False,
        width=120,
    )


def build_merged_parent_map(ws: Any) -> dict[tuple[int, int], tuple[int, int]]:
    parent: dict[tuple[int, int], tuple[int, int]] = {}
    for cell_range in ws.merged_cells.ranges:
        top_left = (cell_range.min_row, cell_range.min_col)
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                if (row, col) != top_left:
                    parent[(row, col)] = top_left
    return parent


def merged_value(ws: Any, parent_map: dict[tuple[int, int], tuple[int, int]], row: int, col: int) -> str:
    source_row, source_col = parent_map.get((row, col), (row, col))
    return normalize_cell(ws.cell(source_row, source_col).value)


def formula_cells(ws: Any) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                cells.append({"coordinate": cell.coordinate, "row": cell.row, "column": cell.column})
    return cells


def probe_read_only_dimensions(path: Path, sheets: list[str] | tuple[str, ...] | None = None) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict[str, Any]] = {}
    wanted = set(sheets) if sheets else set(wb.sheetnames)
    try:
        for ws in wb.worksheets:
            if ws.title not in wanted:
                continue
            before = ws.calculate_dimension()
            reset = False
            if before == "A1:A1":
                ws.reset_dimensions()
                reset = True
            rows = 0
            cols = 0
            for row in ws.iter_rows(values_only=True):
                rows += 1
                cols = max(cols, len(row))
            out[ws.title] = {
                "read_only_dimension_before": before,
                "reset_dimensions_applied": reset,
                "read_only_rows_after_gate": rows,
                "read_only_columns_after_gate": cols,
            }
    finally:
        wb.close()
    return out


@dataclass
class SheetExtract:
    title: str
    headers: list[str]
    rows: list[dict[str, Any]]
    max_row: int
    max_column: int
    calculate_dimension: str
    merged_ranges_count: int
    merged_filled_cells_count: int
    formula_cells_count: int


def extract_workbook(path: Path, sheets: list[str] | tuple[str, ...] | None = None) -> dict[str, SheetExtract]:
    read_only_probe = probe_read_only_dimensions(path, sheets)
    data_wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    extracts: dict[str, SheetExtract] = {}
    try:
        wanted = set(sheets) if sheets else {ws.title for ws in data_wb.worksheets}
        for title in data_wb.sheetnames:
            if title not in wanted:
                continue
            ws = data_wb[title]
            fws = formula_wb[title]
            parent_map = build_merged_parent_map(ws)
            headers = [merged_value(ws, parent_map, 1, col) for col in range(1, ws.max_column + 1)]
            rows: list[dict[str, Any]] = []
            for row_no in range(2, ws.max_row + 1):
                values = [merged_value(ws, parent_map, row_no, col) for col in range(1, ws.max_column + 1)]
                raw_values = [normalize_cell(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
                rows.append(
                    {
                        "row_no": row_no,
                        "values": values,
                        "raw_values": raw_values,
                        "nonblank": any(values),
                        "raw_nonblank": any(raw_values),
                    }
                )
            probe = read_only_probe.get(title, {})
            extracts[title] = SheetExtract(
                title=title,
                headers=headers,
                rows=rows,
                max_row=ws.max_row,
                max_column=ws.max_column,
                calculate_dimension=ws.calculate_dimension(),
                merged_ranges_count=len(ws.merged_cells.ranges),
                merged_filled_cells_count=len(parent_map),
                formula_cells_count=len(formula_cells(fws)),
            )
            extracts[title].__dict__.update(probe)
    finally:
        data_wb.close()
        formula_wb.close()
    return extracts


def header_index(headers: list[str], name: str) -> int:
    if name not in headers:
        raise C1Error(f"missing required header: {name}")
    return headers.index(name)


def row_field(sheet: SheetExtract, row: dict[str, Any], header: str) -> str:
    idx = header_index(sheet.headers, header)
    return row["values"][idx] if idx < len(row["values"]) else ""


def content_digest_for_workbook(path: Path, expected_sheets: list[str]) -> tuple[str, dict[str, Any]]:
    read_only_probe = probe_read_only_dimensions(path, expected_sheets)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    normalized: list[dict[str, Any]] = []
    sheet_stats: dict[str, Any] = {}
    try:
        for title in expected_sheets:
            if title not in wb.sheetnames:
                continue
            ws = wb[title]
            before = ws.calculate_dimension()
            if before == "A1:A1":
                ws.reset_dimensions()
            rows_after_header = 0
            nonblank = 0
            max_cols = 0
            for row_no, values in enumerate(ws.iter_rows(values_only=True), 1):
                row_values = [normalize_cell(value) for value in values]
                max_cols = max(max_cols, len(row_values))
                if row_no == 1:
                    continue
                rows_after_header += 1
                if any(row_values):
                    nonblank += 1
                    normalized.append({"sheet": title, "row_no": row_no, "values": row_values})
            probe = read_only_probe.get(title, {})
            sheet_stats[title] = {
                "rows_after_header": rows_after_header,
                "nonblank_rows_after_header": nonblank,
                "max_row": rows_after_header + 1,
                "max_column": max_cols,
                "calculate_dimension": probe.get("read_only_dimension_before", ""),
                "merged_ranges_count": None,
                "merged_filled_cells_count": None,
                "formula_cells_count": None,
                "read_only_dimension_before": probe.get("read_only_dimension_before", ""),
                "reset_dimensions_applied": probe.get("reset_dimensions_applied", False),
                "read_only_rows_after_gate": probe.get("read_only_rows_after_gate", 0),
                "read_only_columns_after_gate": probe.get("read_only_columns_after_gate", 0),
            }
    finally:
        wb.close()
    digest = sha256_json(normalized)
    missing = [sheet for sheet in expected_sheets if sheet not in sheet_stats]
    return digest, {"sheet_stats": sheet_stats, "missing_expected_sheets": missing}


def manifest_path() -> Path:
    return CONTRACTS_DIR / "source-snapshot-manifest.yaml"


def load_manifest() -> dict[str, Any]:
    path = manifest_path()
    if not path.exists():
        raise C1Error(f"missing manifest: {path}. Run scripts/freeze_snapshot.py first.")
    return safe_load_yaml(path)


def snapshot_file_path(manifest: dict[str, Any], source_key: str) -> Path:
    for source in manifest.get("c1_semantic_sources", []):
        if source.get("source_key") == source_key:
            return Path(source["snapshot_file"])
    for source in manifest.get("c1_reference_sources", []):
        if source.get("source_key") == source_key:
            return Path(source["snapshot_file"])
    raise C1Error(f"source key not found in manifest: {source_key}")


def copy_sources_to_snapshot(snapshot_id: str, sources: list[dict[str, Any]]) -> Path:
    snapshot_dir = RAW_SNAPSHOT_ROOT / snapshot_id
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    for source in sources:
        src = downloads_path(source["filename"])
        if not src.exists():
            raise C1Error(f"missing source xlsx: {src}")
        dst = snapshot_dir / source["filename"]
        if not dst.exists() or file_sha256(dst) != file_sha256(src):
            shutil.copy2(src, dst)
    return snapshot_dir


def parse_ds_protocol(text: str) -> dict[str, Any]:
    return json.loads(text, object_pairs_hook=reject_duplicate_json_keys)


def extract_value_tuple(slots: dict[str, Any], source_sheet: str, source_row_no: int) -> dict[str, str]:
    raw = slots.get("value")
    if raw in (None, ""):
        return {"ref": "", "direct": "", "offset": "", "type": ""}
    if not isinstance(raw, dict):
        raise ValueExtractionError("value is present but is not an object", source_sheet, source_row_no, raw)
    missing = [key for key in ("ref", "direct", "offset", "type") if key not in raw]
    if missing:
        raise ValueExtractionError(
            f"value object missing keys: {','.join(missing)}",
            source_sheet,
            source_row_no,
            raw,
        )
    return {key: normalize_cell(raw.get(key)) for key in ("ref", "direct", "offset", "type")}


def normalize_primitive(action_code: str, intent: str) -> str:
    code = action_code or intent
    if "to_percent" in code or "percent" in code:
        return "by_percent"
    if "increase_value_by_specific_value" in code or "raise_value_by_specific_value" in code:
        return "increase_by_number"
    if "decrease_value_by_specific_value" in code or "lower_value_by_specific_value" in code:
        return "decrease_by_number"
    if "to_specific_value" in code or "to_number" in code or "by_number" in code or "by_specific_value" in code:
        return "adjust_to_number"
    if "by_exp" in code or "little" in code:
        if code.startswith(("lower", "decrease", "close")):
            return "decrease_by_exp"
        return "increase_by_exp"
    if "to_max" in code:
        return "adjust_to_max"
    if "to_min" in code:
        return "adjust_to_min"
    if "to_gear" in code:
        return "adjust_to_gear"
    if "mode" in code or code.startswith(("set_", "switch_")):
        return "set_mode"
    if code.startswith(("open_", "activate_")):
        return "power_on"
    if code.startswith(("close_", "deactivate_")):
        return "power_off"
    if code.startswith(("query_", "check_")):
        return "query"
    if code.startswith("raise_"):
        return "increase_by_exp"
    if code.startswith("lower_"):
        return "decrease_by_exp"
    return re.sub(r"[^a-z0-9_]+", "_", code.lower()).strip("_") or "unknown"


def derive_device(intent: str) -> str:
    core = ACTION_PREFIX.sub("", intent)
    core = VALUE_SUFFIX.sub("", core)
    core = VALUE_SUFFIX.sub("", core)
    return re.sub(r"[^a-z0-9_]+", "_", core.lower()).strip("_") or "unknown"


def slot_identity(slots: dict[str, Any]) -> tuple[str, list[str]]:
    keys = sorted(key for key in slots if key != "value")
    return ("+".join(keys) if keys else "none", keys)


def classify_range(semantic_range: str) -> str:
    if not semantic_range:
        return "none"
    if "<" in semantic_range and ">" in semantic_range:
        return "placeholder_open"
    return "material_candidate"


def failure_receipt_path() -> Path:
    return CONTRACTS_DIR / "semantic-function-contract.failure-receipt.md"


def write_failure_receipt(error: ValueExtractionError) -> None:
    body = "\n".join(
        [
            "# C1 codegen failure receipt",
            "",
            "state: blocked",
            "reason: value_tuple_ambiguity",
            f"source_sheet: {error.source_sheet}",
            f"source_row_no: {error.source_row_no}",
            "",
            "## Error",
            str(error),
            "",
            "## Raw value payload",
            "```json",
            stable_json(error.payload, ensure_ascii=False),
            "```",
            "",
        ]
    )
    atomic_write_text(failure_receipt_path(), body)
